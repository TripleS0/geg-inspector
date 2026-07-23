"""HTTP-level tests against the FastAPI backend using TestClient."""

from __future__ import annotations

import importlib
import io
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch


class BackendApiTests(unittest.TestCase):
    """Smoke-test the most user-visible API surfaces."""

    def setUp(self) -> None:
        try:
            from fastapi.testclient import TestClient  # noqa: F401
        except Exception as err:  # pragma: no cover - fastapi optional
            self.skipTest(f"fastapi/starlette not installed: {err}")

        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self._old_db = os.environ.get("DATAFUSIONX_DB_PATH")
        self._old_home = os.environ.get("DATAFUSIONX_HOME")
        os.environ["DATAFUSIONX_DB_PATH"] = str(Path(self._tmp.name) / "api.sqlite3")
        os.environ["DATAFUSIONX_HOME"] = self._tmp.name

        from app import runtime_paths as rp
        importlib.reload(rp)
        from backend import main as backend_main
        importlib.reload(backend_main)
        from fastapi.testclient import TestClient

        self.client = TestClient(backend_main.app)
        login = self.client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin123"},
        )
        self.assertEqual(login.status_code, 200, login.text)
        token = login.json()["access_token"]
        self.client.headers.update({"Authorization": f"Bearer {token}"})

    def tearDown(self) -> None:
        if self._old_db is None:
            os.environ.pop("DATAFUSIONX_DB_PATH", None)
        else:
            os.environ["DATAFUSIONX_DB_PATH"] = self._old_db
        if self._old_home is None:
            os.environ.pop("DATAFUSIONX_HOME", None)
        else:
            os.environ["DATAFUSIONX_HOME"] = self._old_home
        from app import runtime_paths as rp
        importlib.reload(rp)
        from backend import main as backend_main
        importlib.reload(backend_main)

    def test_health(self) -> None:
        resp = self.client.get("/api/health")
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertEqual(body["status"], "ok")
        self.assertTrue(body["db_path"].endswith("api.sqlite3"))

    def test_batches_when_empty(self) -> None:
        resp = self.client.get("/api/batches")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), {"items": []})

    def test_tables_when_empty(self) -> None:
        resp = self.client.get("/api/tables")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), {"items": []})

    def test_import_validation(self) -> None:
        resp = self.client.post("/api/import/bank", json={"file_paths": [], "bank_name": "BankA"})
        self.assertEqual(resp.status_code, 400)

    def test_import_unknown_source(self) -> None:
        resp = self.client.post("/api/import/foo", json={"file_paths": ["a.xlsx"], "bank_name": "BankA"})
        self.assertEqual(resp.status_code, 400)

    def test_task_not_found(self) -> None:
        resp = self.client.get("/api/tasks/non-existing")
        self.assertEqual(resp.status_code, 404)

    def test_qichacha_export_requires_credentials(self) -> None:
        old_k = os.environ.pop("QICHACHA_APP_KEY", None)
        old_s = os.environ.pop("QICHACHA_SECRET_KEY", None)
        try:
            resp = self.client.post("/api/qichacha/basic-details/export", data={"keywords": "测试公司"})
            self.assertEqual(resp.status_code, 503)
        finally:
            if old_k is not None:
                os.environ["QICHACHA_APP_KEY"] = old_k
            if old_s is not None:
                os.environ["QICHACHA_SECRET_KEY"] = old_s

    def test_qichacha_export_mocked_and_logs(self) -> None:
        os.environ["QICHACHA_APP_KEY"] = "unit_test_key"
        os.environ["QICHACHA_SECRET_KEY"] = "unit_test_secret"
        sample = {
            "Status": "200",
            "Message": "【有效请求】查询成功",
            "OrderNumber": "ECI2026050918570871609609",
            "Result": {
                "Name": "广州大泽森工贸有限公司",
                "CreditCode": "91440101749901601M",
                "Area": {"Province": "广东省", "City": "广州市", "County": "天河区"},
                "OriginalName": [],
            },
        }
        with patch("backend.main.fetch_basic_details_by_name", return_value=sample):
            q = self.client.post("/api/qichacha/basic-details/query", data={"keywords": "广州大泽森工贸有限公司"})
        self.assertEqual(q.status_code, 200, q.text)
        body = q.json()
        self.assertEqual(body["count"], 1)
        self.assertEqual(len(body["rows"]), 1)

        resp = self.client.post(
            "/api/qichacha/basic-details/export",
            json={"rows": body["rows"], "run_id": body["run_id"]},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        self.assertIn("spreadsheet", resp.headers.get("content-type", ""))
        self.assertEqual(resp.headers.get("X-Run-Id"), body["run_id"])

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(resp.content))
        self.assertEqual(wb.sheetnames[0], "工商信息")
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Name", headers)
        self.assertIn("query_keyword", headers)

        log_resp = self.client.get("/api/qichacha/query-logs?limit=10")
        self.assertEqual(log_resp.status_code, 200)
        items = log_resp.json()["items"]
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["query_keyword"], "广州大泽森工贸有限公司")
        self.assertEqual(items[0]["matched_name"], "广州大泽森工贸有限公司")
        self.assertEqual(items[0]["credit_code"], "91440101749901601M")

    def test_qichacha_ingest_profile_and_risk_rules(self) -> None:
        ingest = self.client.post(
            "/api/qichacha/ingest-profile",
            json={
                "rows": [
                    {
                        "Name": "测试工商主体有限公司",
                        "CreditCode": "91110000MA01234567",
                        "OperName": "张三",
                    }
                ],
                "run_id": "test-run",
            },
        )
        self.assertEqual(ingest.status_code, 200, ingest.text)
        ing = ingest.json()
        self.assertIn("import_batch_id", ing)
        self.assertGreaterEqual(ing.get("rows_total", 0), 1)

        rules = self.client.get("/api/commercial/risk-rules")
        self.assertEqual(rules.status_code, 200, rules.text)
        codes = {item["rule_code"] for item in rules.json()["items"]}
        self.assertIn("R007", codes)
        r007 = next(r for r in rules.json()["items"] if r["rule_code"] == "R007")
        params = r007["params"]
        patch = self.client.patch(
            "/api/commercial/risk-rules/R007",
            json={"params": {**params, "min_shared_inquiries": 4}},
        )
        self.assertEqual(patch.status_code, 200, patch.text)
        self.assertEqual(patch.json()["params"]["min_shared_inquiries"], 4)

        em = self.client.get(
            "/api/commercial/00000000-0000-0000-0000-000000000001/entity-matches?limit=10"
        )
        self.assertEqual(em.status_code, 200, em.text)
        self.assertIn("items", em.json())

    def test_batches_merged_includes_enterprise_and_delete(self) -> None:
        ingest = self.client.post(
            "/api/qichacha/ingest-profile",
            json={
                "rows": [
                    {"Name": "批次合并测试公司", "CreditCode": "91110000MA09999999", "OperName": "李四"},
                ],
            },
        )
        self.assertEqual(ingest.status_code, 200, ingest.text)
        batch_id = ingest.json()["import_batch_id"]

        all_resp = self.client.get("/api/batches?limit=200")
        self.assertEqual(all_resp.status_code, 200, all_resp.text)
        rows = all_resp.json()["items"]
        hit = next((r for r in rows if r["import_batch_id"] == batch_id), None)
        self.assertIsNotNone(hit)
        self.assertEqual(hit["source_type"], "enterprise")

        del_resp = self.client.delete(f"/api/batches/{batch_id}")
        self.assertEqual(del_resp.status_code, 200, del_resp.text)
        body = del_resp.json()
        self.assertEqual(body.get("status"), "ok")
        self.assertEqual(body.get("source_type"), "enterprise")

        all_resp2 = self.client.get("/api/batches?limit=200")
        ids2 = {r["import_batch_id"] for r in all_resp2.json()["items"]}
        self.assertNotIn(batch_id, ids2)

        bad = self.client.delete("/api/batches/__nonexistent_batch__")
        self.assertEqual(bad.status_code, 400)


if __name__ == "__main__":
    unittest.main()
