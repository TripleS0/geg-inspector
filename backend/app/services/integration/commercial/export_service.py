"""Commercial-source export service with stable column layout."""

from __future__ import annotations

import re
from typing import Any

from app.services.integration.bank.export_service import BankExportService


def _strip_company_q_suffix(text: str) -> str:
    """Remove trailing inquiry-code suffix like 公司名(Q231201...) from display names."""
    value = (text or "").strip()
    if not value:
        return ""
    value = re.sub(r"\s*[\(（]\s*Q[A-Za-z0-9._\-]+(?:\s*[)）])?\s*$", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+Q[A-Za-z0-9._\-]+\s*$", "", value, flags=re.IGNORECASE)
    return value.strip(" \t()（）[]【】")


def _split_company_names(text: str) -> list[str]:
    raw = (text or "").strip()
    if not raw:
        return []
    raw = re.sub(r"(中标供应商|供应商|公司名称)\s*[:：]", "", raw)
    parts = re.split(r"[\r\n、,，;；/／|｜]+", raw)
    out: list[str] = []
    seen: set[str] = set()
    for part in parts:
        name = _strip_company_q_suffix(part)
        if not name or name in {"无", "未公布", "无中标", "流标", "废标"}:
            continue
        key = re.sub(r"\s+", "", name).lower()
        if key in seen:
            continue
        out.append(name)
        seen.add(key)
    return out


class CommercialExportService(BankExportService):
    """Export commercial integration result using fixed sample-aligned columns."""

    @staticmethod
    def export_basis_description() -> str:
        """User-facing note for commercial export (two sheets when summary is available)."""
        return (
            "导出内容：当前窗口中「最后一次整合入库」对应批次的商务网整合结果。"
            "工作表①「全字段合并」：固定模板列，首列为“数据来源”。"
            "工作表②「中标情况统计」：本批次中标/未中标等汇总（有汇总数据时写入）。"
        )

    OUTPUT_COLUMNS = [
        "数据来源",
        "询价单号",
        "摘要",
        "采购单位",
        "寻源策略",
        "询价类别",
        "预估去税总额",
        "公布中标信息",
        "提交人/时间",
        "报价截止时间",
        "报价方式",
        "联系人手机",
        "联系人邮箱",
        "状态",
        "询价类型",
        "预估含税总额",
        "寻源范围",
        "要求到货日期",
        "质疑截止时间",
        "报价类别",
        "固定电话",
        "货币",
        "中标原因",
        "中标供应商",
        "备注",
        "序号",
        "物资编码/来源采购申请代码--物资描述",
        "型号规格",
        "品牌/厂家/产地",
        "补充说明",
        "单位",
        "数量",
        "预估单价 (含税)",
        "预估总价 (含税)",
        "公司名称",
        "含税单价(元)",
        "总价(元)",
        "税率",
        "产地",
        "品牌",
        "交货日期",
        "供应商备注",
        "不含税合计总价(元)",
        "税额(元)",
        "含税合计总价(元)",
        "中标金额(元)",
        "中标总金额（元）",
    ]

    FIELD_ALIASES = {
        "物资编码/来源采购申请代码--物资描述": {
            "物资编码/来源采购申请代码--物资描述",
            "物资编码/来源采购申请代码-物资描述",
            "物资编码_来源采购申请代码--物资描述",
        },
        "税额(元)": {"税额(元)", "税金(元)"},
        "预估含税总额": {"预估含税总额", "预估含税总额(元)"},
        "预估去税总额": {"预估去税总额", "预估去税总额(元)"},
    }

    BASE_MERGE_COLUMNS = [
        "数据来源",
        "询价单号",
        "摘要",
        "采购单位",
        "寻源策略",
        "询价类别",
        "预估去税总额",
        "公布中标信息",
        "提交人/时间",
        "报价截止时间",
        "报价方式",
        "联系人手机",
        "联系人邮箱",
        "状态",
        "询价类型",
        "预估含税总额",
        "寻源范围",
        "要求到货日期",
        "质疑截止时间",
        "报价类别",
        "固定电话",
        "货币",
        "中标原因",
        "中标供应商",
        "备注",
        "中标总金额（元）",
    ]
    SUPPLIER_GROUP_MERGE_COLUMNS = [
        "公司名称",
        "不含税合计总价(元)",
        "税额(元)",
        "含税合计总价(元)",
        "中标金额(元)",
    ]
    BID_SUMMARY_SHEET = "中标情况统计"
    BID_SUMMARY_COLUMNS = [
        "序号",
        "公司名称",
        "中标次数",
        "中标总金额",
        "不中标次数",
        "采购单位",
        "公司中标次数",
        "公司中标金额",
        "中标金额",
        "询价单号",
        "备注",
    ]

    def export_batch_to_xlsx(self, import_batch_id: str, output_path: str) -> str:
        """Export with merged-cell layout matching template style."""
        try:
            import pandas as pd
        except ImportError as err:
            raise RuntimeError("缺少 pandas，请先执行: pip install -r requirements.txt") from err

        out_file = self._normalize_output_xlsx_path(output_path)
        all_headers, all_rows = self._load_batch_all_raw_fields(import_batch_id)
        df_all = pd.DataFrame(all_rows, columns=all_headers)
        if all_headers == self.OUTPUT_COLUMNS:
            df_all = df_all.reindex(columns=self.OUTPUT_COLUMNS, fill_value="")

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            df_all.to_excel(writer, index=False, sheet_name="全字段合并")
            ws = writer.sheets.get("全字段合并")
            if ws is not None:
                if all_headers == self.OUTPUT_COLUMNS and all_rows:
                    self._apply_template_merges(ws, all_headers, all_rows)
                self._apply_sheet_style(ws)

            if all_headers == self.OUTPUT_COLUMNS and all_rows:
                summary_rows = self._build_bid_summary_rows(all_headers, all_rows)
                if summary_rows:
                    summary_df = pd.DataFrame(summary_rows, columns=self.BID_SUMMARY_COLUMNS)
                    summary_df.to_excel(writer, index=False, sheet_name=self.BID_SUMMARY_SHEET)
                    summary_ws = writer.sheets.get(self.BID_SUMMARY_SHEET)
                    if summary_ws is not None:
                        self._apply_summary_merges(summary_ws)
                        self._apply_sheet_style(summary_ws)
        return str(out_file)

    def _load_batch_all_raw_fields(self, import_batch_id: str) -> tuple[list[str], list[list[str]]]:
        """Load commercial rows in fixed layout; fallback to generic union export."""
        rows = self._load_commercial_rows(import_batch_id)
        if not rows:
            return super()._load_batch_all_raw_fields(import_batch_id)
        rows = self._format_rows_like_template(rows)
        self._fill_bid_total_per_inquiry(rows)
        output_rows = []
        for row in rows:
            output_rows.append([self._to_text(row.get(col, "")) for col in self.OUTPUT_COLUMNS])
        return self.OUTPUT_COLUMNS, output_rows

    def _load_commercial_rows(self, import_batch_id: str) -> list[dict[str, str]]:
        file_rows = self._client.query_all(
            """
            SELECT file_id, file_name
            FROM meta_bank_files
            WHERE import_batch_id=? AND source_type='commercial';
            """,
            (import_batch_id,),
        )
        file_name_map = {int(row[0]): str(row[1]) for row in file_rows if row and row[0] is not None}
        if not file_name_map:
            return []

        sheet_rows = self._client.query_all(
            """
            SELECT DISTINCT s.raw_table_name
            FROM meta_bank_sheets s
            JOIN meta_bank_files f ON f.file_id=s.file_id
            WHERE f.import_batch_id=? AND f.source_type='commercial'
            ORDER BY s.raw_table_name;
            """,
            (import_batch_id,),
        )
        table_names = [str(row[0]) for row in sheet_rows if row and row[0]]
        if not table_names:
            return []

        output: list[dict[str, str]] = []
        for table in table_names:
            info = self._client.query_all(f"PRAGMA table_info({self._client.quote_ident(table)});")
            src_cols = [str(x[1]) for x in info if str(x[1]).startswith("src_")]
            if not src_cols:
                continue
            sql_cols = ", ".join(self._client.quote_ident(c) for c in src_cols)
            raw_rows = self._client.query_all(
                f"""
                SELECT source_file_id, source_sheet, {sql_cols}
                FROM {self._client.quote_ident(table)}
                WHERE import_batch_id=?
                ORDER BY raw_id;
                """,
                (import_batch_id,),
            )
            label_map = {self._display_src_name(c): idx for idx, c in enumerate(src_cols, start=2)}
            normalized_label_map = {self._normalize_label(k): v for k, v in label_map.items()}
            for row in raw_rows:
                source_file_id = int(row[0]) if row[0] is not None else 0
                file_name = file_name_map.get(source_file_id, "")
                source_sheet = self._to_text(row[1])
                record: dict[str, str] = {
                    "数据来源": self._build_row_source_name(file_name, source_sheet, fallback_table=table)
                }
                for out_col in self.OUTPUT_COLUMNS[1:]:
                    if out_col == "中标总金额（元）":
                        record[out_col] = ""
                        continue
                    idx = self._resolve_index(label_map, normalized_label_map, out_col)
                    record[out_col] = "" if idx is None else self._to_text(row[idx])
                output.append(record)
        return output

    def _fill_bid_total_per_inquiry(self, rows: list[dict[str, str]]) -> None:
        """同一询价单内：按公司名称各取一次中标金额(元)后相加（避免多行明细重复累计）。"""
        from collections import defaultdict

        grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in rows:
            key = self._to_text(row.get("询价单号", "")) or self._to_text(row.get("数据来源", ""))
            grouped[key].append(row)

        totals: dict[str, str] = {}
        for key, group in grouped.items():
            if not group:
                totals[key] = ""
                continue
            current_company = ""
            company_amount: dict[str, float] = {}
            for r in group:
                co = self._to_text(r.get("公司名称", ""))
                if co:
                    current_company = co
                if not current_company:
                    continue
                if current_company in company_amount:
                    continue
                company_amount[current_company] = self._safe_amount(
                    self._to_text(r.get("中标金额(元)", ""))
                )
            total_val = sum(company_amount.values())
            totals[key] = self._format_amount_cell(total_val)

        for row in rows:
            key = self._to_text(row.get("询价单号", "")) or self._to_text(row.get("数据来源", ""))
            row["中标总金额（元）"] = totals.get(key, "")

    def _format_amount_cell(self, value: float) -> str:
        """Format numeric total for Excel cell text."""
        rounded = round(float(value), 2)
        if rounded == int(rounded):
            return str(int(rounded))
        return str(rounded)

    def _format_rows_like_template(self, rows: list[dict[str, str]]) -> list[dict[str, str]]:
        """Reorder and blank repeated cells to match expected sample style."""
        source_rank: dict[str, int] = {}
        supplier_rank: dict[tuple[str, str], int] = {}
        for idx, row in enumerate(rows):
            source = self._to_text(row.get("数据来源", ""))
            company = self._to_text(row.get("公司名称", ""))
            if source not in source_rank:
                source_rank[source] = len(source_rank)
            key = (source, company)
            if company and key not in supplier_rank:
                supplier_rank[key] = len([k for k in supplier_rank.keys() if k[0] == source])

        indexed = list(enumerate(rows))
        ordered = [
            row
            for _, row in sorted(
                indexed,
                key=lambda p: (
                    source_rank.get(self._to_text(p[1].get("数据来源", "")), 10**9),
                    supplier_rank.get(
                        (self._to_text(p[1].get("数据来源", "")), self._to_text(p[1].get("公司名称", ""))),
                        10**9,
                    ),
                    self._sequence_key(p[1].get("序号", "")),
                    p[0],
                ),
            )
        ]
        last_source = ""
        last_group = ("", "")
        for row in ordered:
            source = self._to_text(row.get("数据来源", ""))
            company = self._to_text(row.get("公司名称", ""))
            if source == last_source:
                row["数据来源"] = ""
            else:
                last_source = source
            group = (source, company)
            if group == last_group:
                row["公司名称"] = ""
            else:
                last_group = group
        return ordered

    def _sequence_key(self, value: Any) -> tuple[int, str]:
        text = self._to_text(value)
        if not text:
            return (10**9, "")
        try:
            return (int(float(text)), text)
        except Exception:
            return (10**9, text)

    def _resolve_index(
        self, label_map: dict[str, int], normalized_label_map: dict[str, int], output_col: str
    ) -> int | None:
        aliases = self.FIELD_ALIASES.get(output_col, {output_col})
        for alias in aliases:
            idx = label_map.get(alias)
            if idx is not None:
                return idx
            normalized_idx = normalized_label_map.get(self._normalize_label(alias))
            if normalized_idx is not None:
                return normalized_idx
        return None

    def _normalize_label(self, value: str) -> str:
        text = self._to_text(value)
        return "".join(ch for ch in text if ch.isalnum())

    def _to_text(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text.lower() == "nan":
            return ""
        return text

    def _normalize_output_xlsx_path(self, output_path: str):
        from pathlib import Path

        out_file = Path(output_path)
        if out_file.suffix.lower() != ".xlsx":
            out_file = out_file.with_suffix(".xlsx")
        out_file.parent.mkdir(parents=True, exist_ok=True)
        return out_file

    def _apply_template_merges(self, ws, headers: list[str], rows: list[list[str]]) -> None:
        """Merge base columns by source and merge supplier name by group."""
        col_idx = {name: i + 1 for i, name in enumerate(headers)}
        source_col = "数据来源"
        company_col = "公司名称"
        if source_col not in col_idx or company_col not in col_idx:
            return

        effective_source: list[str] = []
        effective_company: list[str] = []
        current_source = ""
        current_company = ""
        for row in rows:
            row_map = {headers[i]: self._to_text(row[i]) for i in range(min(len(headers), len(row)))}
            source = row_map.get(source_col, "")
            company = row_map.get(company_col, "")
            if source:
                current_source = source
                current_company = ""
            if company:
                current_company = company
            effective_source.append(current_source)
            effective_company.append(current_company)

        # 1) Base fields merged by source block.
        start = 0
        while start < len(rows):
            src = effective_source[start]
            end = start
            while end + 1 < len(rows) and effective_source[end + 1] == src:
                end += 1
            if src and end > start:
                for name in self.BASE_MERGE_COLUMNS:
                    idx = col_idx.get(name)
                    if idx is None:
                        continue
                    ws.merge_cells(start_row=start + 2, start_column=idx, end_row=end + 2, end_column=idx)
            start = end + 1

        # 2) Supplier-group fields merged by (source, company) contiguous block.
        start = 0
        while start < len(rows):
            key = (effective_source[start], effective_company[start])
            end = start
            while end + 1 < len(rows) and (effective_source[end + 1], effective_company[end + 1]) == key:
                end += 1
            if key[1] and end > start:
                for name in self.SUPPLIER_GROUP_MERGE_COLUMNS:
                    idx = col_idx.get(name)
                    if idx is None:
                        continue
                    ws.merge_cells(start_row=start + 2, start_column=idx, end_row=end + 2, end_column=idx)
            start = end + 1

    def _apply_sheet_style(self, ws) -> None:
        """Apply borders, alignment, freeze header row, and autosize columns."""
        from openpyxl.styles import Alignment, Border, Side

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(vertical="center", wrap_text=True)

        # Freeze first row.
        ws.freeze_panes = "A2"

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = align
                cell.border = border

        # Auto-fit each column width based on displayed content length.
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                text = "" if cell.value is None else str(cell.value)
                if len(text) > max_len:
                    max_len = len(text)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 80)

    def _build_bid_summary_rows(self, headers: list[str], rows: list[list[str]]) -> list[list[Any]]:
        """Build bid-summary rows with grain: company + purchaser + inquiry no."""
        required = {"公司名称", "中标供应商", "中标金额(元)", "采购单位", "询价单号", "数据来源"}
        col_map = {name: idx for idx, name in enumerate(headers)}
        if any(name not in col_map for name in required):
            return []

        event_map: dict[tuple[str, str, str], dict[str, Any]] = {}
        current_source = ""
        current_company = ""
        for raw_row in rows:
            source_text = self._to_text(raw_row[col_map["数据来源"]]) if col_map["数据来源"] < len(raw_row) else ""
            company_text = self._to_text(raw_row[col_map["公司名称"]]) if col_map["公司名称"] < len(raw_row) else ""
            if source_text:
                current_source = source_text
                current_company = ""
            if company_text:
                current_company = company_text
            effective_companies = _split_company_names(current_company)
            if not effective_companies:
                continue

            purchaser = self._to_text(raw_row[col_map["采购单位"]]) if col_map["采购单位"] < len(raw_row) else ""
            inquiry_no = self._to_text(raw_row[col_map["询价单号"]]) if col_map["询价单号"] < len(raw_row) else ""
            winner = self._to_text(raw_row[col_map["中标供应商"]]) if col_map["中标供应商"] < len(raw_row) else ""
            amount_text = self._to_text(raw_row[col_map["中标金额(元)"]]) if col_map["中标金额(元)"] < len(raw_row) else ""
            winner_norms = [
                self._normalize_company_name(name)
                for name in _split_company_names(winner)
            ]

            for effective_company in effective_companies:
                company_norm = self._normalize_company_name(effective_company)
                if not company_norm:
                    continue
                key = (company_norm, purchaser, inquiry_no)
                event = event_map.setdefault(
                    key,
                    {
                        "公司名称": effective_company,
                        "采购单位": purchaser,
                        "询价单号": inquiry_no,
                        "winners": set(),
                        "has_winner": False,
                        "win_amount_candidates": set(),
                        "source": current_source,
                    },
                )
                # Keep the first non-empty display name for better readability.
                if not event["公司名称"] and effective_company:
                    event["公司名称"] = effective_company
                if not event["采购单位"] and purchaser:
                    event["采购单位"] = purchaser
                if not event["询价单号"] and inquiry_no:
                    event["询价单号"] = inquiry_no

                for winner_norm in winner_norms:
                    if winner_norm:
                        event["winners"].add(winner_norm)
                        event["has_winner"] = True
                        if self._is_company_winner(company_norm, winner_norm):
                            amount_value = self._safe_amount(amount_text)
                            if amount_value > 0:
                                event["win_amount_candidates"].add(amount_value)

        if not event_map:
            return []

        base_rows: list[dict[str, Any]] = []
        for _, event in event_map.items():
            company_norm = self._normalize_company_name(self._to_text(event["公司名称"]))
            has_winner = bool(event["has_winner"])
            winners: set[str] = event["winners"]
            is_win = bool(company_norm) and any(self._is_company_winner(company_norm, w) for w in winners)
            win_count = 1 if is_win else 0
            lose_count = 1 if has_winner and not is_win else 0
            # 同一询价单通常有多条明细行，金额字段会重复出现；这里对候选值去重后只取一次。
            candidates: set[float] = event["win_amount_candidates"]
            amount = max(candidates) if (is_win and candidates) else 0.0
            base_rows.append(
                {
                    "公司名称": self._to_text(event["公司名称"]),
                    "公司标准名": company_norm,
                    "采购单位": self._to_text(event["采购单位"]),
                    "询价单号": self._to_text(event["询价单号"]),
                    "中标次数": win_count,
                    "金额": amount,
                    "不中标次数": lose_count,
                    "备注": "",
                }
            )

        company_org_totals: dict[tuple[str, str], dict[str, float]] = {}
        company_totals: dict[str, dict[str, float]] = {}
        for row in base_rows:
            company_key = self._to_text(row["公司标准名"])
            key = (company_key, self._to_text(row["采购单位"]))
            item = company_org_totals.setdefault(key, {"中标次数": 0.0, "公司中标金额": 0.0})
            item["中标次数"] += float(row["中标次数"])
            item["公司中标金额"] += float(row["金额"])
            company_item = company_totals.setdefault(company_key, {"中标次数": 0.0, "中标金额": 0.0, "不中标次数": 0.0})
            company_item["中标次数"] += float(row["中标次数"])
            company_item["中标金额"] += float(row["金额"])
            company_item["不中标次数"] += float(row["不中标次数"])

        base_rows.sort(
            key=lambda x: (
                -int(company_totals.get(self._to_text(x["公司标准名"]), {}).get("中标次数", 0.0)),
                -int(company_totals.get(self._to_text(x["公司标准名"]), {}).get("不中标次数", 0.0)),
                self._to_text(x["公司标准名"]),
                self._to_text(x["采购单位"]),
                self._to_text(x["询价单号"]),
            )
        )
        out_rows: list[list[Any]] = []
        seq = 0
        last_company = ""
        for row in base_rows:
            company_key = self._to_text(row["公司标准名"])
            key = (company_key, self._to_text(row["采购单位"]))
            totals = company_org_totals.get(key, {"中标次数": 0.0, "公司中标金额": 0.0})
            company_total = company_totals.get(company_key, {"中标次数": 0.0, "中标金额": 0.0, "不中标次数": 0.0})
            if company_key != last_company:
                seq += 1
                last_company = company_key
            out_rows.append(
                [
                    seq,
                    self._to_text(row["公司名称"]),
                    int(company_total["中标次数"]),
                    round(float(company_total["中标金额"]), 2),
                    int(company_total["不中标次数"]),
                    self._to_text(row["采购单位"]),
                    int(totals["中标次数"]),
                    round(float(totals["公司中标金额"]), 2),
                    round(float(row["金额"]), 2),
                    self._to_text(row["询价单号"]),
                    "",
                ]
            )
        return out_rows

    def _normalize_company_name(self, text: str) -> str:
        """Normalize company name for fuzzy identity matching."""
        value = _strip_company_q_suffix(self._to_text(text))
        if not value:
            return ""
        # remove any bracketed code suffix/prefix, e.g. 公司名(Q123), 公司名（Q123）
        value = re.sub(r"[\(（][^)\）]*[\)）]", "", value)
        value = re.sub(r"\s+", "", value).lower()
        return value

    def _safe_amount(self, text: str) -> float:
        value = self._to_text(text).replace(",", "")
        if not value:
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    def _is_company_winner(self, company_norm: str, winner_norm: str) -> bool:
        """Loose match for company and winner after normalization."""
        c = self._to_text(company_norm)
        w = self._to_text(winner_norm)
        if not c or not w:
            return False
        return c == w or c in w or w in c

    def _apply_summary_merges(self, ws) -> None:
        """Apply summary-sheet merges by company and company+purchaser groups."""
        if ws.max_row <= 2:
            return
        headers = [self._to_text(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
        col_idx = {name: i + 1 for i, name in enumerate(headers)}
        seq_col = col_idx.get("序号")
        company_col = col_idx.get("公司名称")
        win_col = col_idx.get("中标次数")
        amount_col = col_idx.get("中标总金额")
        lose_col = col_idx.get("不中标次数")
        org_col = col_idx.get("采购单位")
        comp_win_col = col_idx.get("公司中标次数")
        comp_amt_col = col_idx.get("公司中标金额")
        remark_col = col_idx.get("备注")
        if company_col is None or org_col is None:
            return

        # Merge company column by contiguous same-company block.
        start = 2
        while start <= ws.max_row:
            company = self._normalize_company_name(self._to_text(ws.cell(row=start, column=company_col).value))
            end = start
            while end + 1 <= ws.max_row and self._normalize_company_name(
                self._to_text(ws.cell(row=end + 1, column=company_col).value)
            ) == company:
                end += 1
            if company and end > start:
                if seq_col is not None:
                    ws.merge_cells(start_row=start, start_column=seq_col, end_row=end, end_column=seq_col)
                ws.merge_cells(start_row=start, start_column=company_col, end_row=end, end_column=company_col)
                if win_col is not None:
                    ws.merge_cells(start_row=start, start_column=win_col, end_row=end, end_column=win_col)
                if amount_col is not None:
                    ws.merge_cells(start_row=start, start_column=amount_col, end_row=end, end_column=amount_col)
                if lose_col is not None:
                    ws.merge_cells(start_row=start, start_column=lose_col, end_row=end, end_column=lose_col)
                if remark_col is not None:
                    ws.merge_cells(start_row=start, start_column=remark_col, end_row=end, end_column=remark_col)
            # Merge purchaser and company totals by company+purchaser block.
            sub_start = start
            while sub_start <= end:
                org = self._to_text(ws.cell(row=sub_start, column=org_col).value)
                sub_end = sub_start
                while sub_end + 1 <= end and self._to_text(ws.cell(row=sub_end + 1, column=org_col).value) == org:
                    sub_end += 1
                if org and sub_end > sub_start:
                    ws.merge_cells(start_row=sub_start, start_column=org_col, end_row=sub_end, end_column=org_col)
                    if comp_win_col is not None:
                        ws.merge_cells(
                            start_row=sub_start,
                            start_column=comp_win_col,
                            end_row=sub_end,
                            end_column=comp_win_col,
                        )
                    if comp_amt_col is not None:
                        ws.merge_cells(
                            start_row=sub_start,
                            start_column=comp_amt_col,
                            end_row=sub_end,
                            end_column=comp_amt_col,
                        )
                sub_start = sub_end + 1
            start = end + 1


__all__ = ["CommercialExportService"]

