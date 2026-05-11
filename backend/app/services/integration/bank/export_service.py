"""Export service for merged raw fields (user-facing Excel)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from app.services.integration.bank.analysis_modules import (
    AnalysisModuleId,
    ModuleParams,
    aggregate_large_flow,
    build_custom_filter_analysis_remark,
    build_amount_rounded_counts,
    filter_large_transactions,
    run_module,
)
from app.services.integration.bank.query_service import BankQueryFilters, BankQueryService
from app.services.shared.db.sqlite_client import SqliteClient

class BankExportService:
    """导出银行整合结果（总体+细则+个人明细+个人统计）。"""
    DISCLAIMER_KEYWORDS = (
        "综合查控平台导出",
        "非实时数据",
        "实际内容以我行生产系统为准",
    )

    def __init__(self, client: SqliteClient | None = None) -> None:
        """Initialize export service."""
        self._client = client or SqliteClient()
        self._query_service = BankQueryService(self._client)

    @staticmethod
    def export_basis_description() -> str:
        """向用户说明导出依据（与界面提示一致）。"""
        return (
            "导出内容：当前窗口中「最后一次整合入库」对应批次的银行整合结果。"
            "工作表①「全字段合并」：总体明细（含数据来源、银行、交易行号/名、备注）。"
            "工作表②「细则_双方往来」：A->B、B->A与差额明细。"
            "工作表③「个人_银行明细」：个人在各银行的逐笔流水。"
            "工作表④「个人_银行统计」：个人在各银行的汇总统计。"
        )

    def export_batch_to_xlsx(self, import_batch_id: str, output_path: str) -> str:
        """将指定 import_batch_id 的全字段合并结果导出为 .xlsx。"""
        try:
            import pandas as pd
        except ImportError as err:
            raise RuntimeError("缺少 pandas，请先执行: pip install -r requirements.txt") from err

        out_file = Path(output_path)
        if out_file.suffix.lower() != ".xlsx":
            out_file = out_file.with_suffix(".xlsx")
        out_file.parent.mkdir(parents=True, exist_ok=True)

        all_headers, all_rows = self._load_batch_all_raw_fields(import_batch_id)
        df_all = pd.DataFrame(all_rows, columns=all_headers)
        std_records = self._load_std_records(import_batch_id)
        detail_headers, detail_rows = self._build_counterparty_detail_sheet(std_records)
        person_detail_headers, person_detail_rows = self._build_person_bank_detail_sheet(std_records)
        person_stat_headers, person_stat_rows = self._build_person_bank_stat_sheet(std_records)
        diag_headers, diag_rows = self._build_name_match_diagnosis_sheet(import_batch_id)

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            df_all.to_excel(writer, index=False, sheet_name="全字段合并")
            ws = writer.sheets.get("全字段合并")
            if ws is not None:
                self._merge_same_source_cells(ws)
                self._apply_sheet_style(ws)
            if detail_rows:
                pd.DataFrame(detail_rows, columns=detail_headers).to_excel(
                    writer, index=False, sheet_name="细则_双方往来"
                )
                detail_ws = writer.sheets.get("细则_双方往来")
                if detail_ws is not None:
                    self._apply_sheet_style(detail_ws)
            if person_detail_rows:
                pd.DataFrame(person_detail_rows, columns=person_detail_headers).to_excel(
                    writer, index=False, sheet_name="个人_银行明细"
                )
                person_detail_ws = writer.sheets.get("个人_银行明细")
                if person_detail_ws is not None:
                    self._apply_sheet_style(person_detail_ws)
            if person_stat_rows:
                pd.DataFrame(person_stat_rows, columns=person_stat_headers).to_excel(
                    writer, index=False, sheet_name="个人_银行统计"
                )
                person_stat_ws = writer.sheets.get("个人_银行统计")
                if person_stat_ws is not None:
                    self._apply_sheet_style(person_stat_ws)
            if diag_rows:
                pd.DataFrame(diag_rows, columns=diag_headers).to_excel(
                    writer, index=False, sheet_name="姓名匹配诊断"
                )
                diag_ws = writer.sheets.get("姓名匹配诊断")
                if diag_ws is not None:
                    self._apply_sheet_style(diag_ws)
        return str(out_file)

    def export_module_report(
        self,
        import_batch_id: str,
        output_path: str,
        module_id: str,
        params: ModuleParams | None = None,
    ) -> str:
        """Export fixed-module analysis: detail sheet, summary sheet, optional flow pivot."""
        try:
            import pandas as pd
        except ImportError as err:
            raise RuntimeError("缺少 pandas，请先执行: pip install -r requirements.txt") from err

        params = params or ModuleParams()
        result = run_module(import_batch_id, module_id, params, self._client)

        out_file = Path(output_path)
        if out_file.suffix.lower() != ".xlsx":
            out_file = out_file.with_suffix(".xlsx")
        out_file.parent.mkdir(parents=True, exist_ok=True)

        detail_headers = [
            "数据来源",
            "银行类别",
            "姓名",
            "卡号",
            "时间戳",
            "收支标志",
            "币种",
            "金额",
            "余额",
            "对手名",
            "对手卡号",
            "交易描述",
            "备注",
        ]
        keys = (
            "data_source",
            "bank_type",
            "person_name",
            "acct_no",
            "txn_time",
            "txn_direction",
            "currency",
            "amount",
            "balance",
            "counterparty_name",
            "counterparty_account",
            "txn_desc",
            "remark",
        )
        detail_rows = [[row.get(k, "") for k in keys] for row in result.hit_records]

        summary = result.summary
        summary_rows: list[list[str]] = [
            ["模块", str(module_id)],
            ["分析说明", result.description],
            ["命中笔数", str(summary.get("txn_count", 0))],
            ["收入总额", self._fmt_amount(float(summary.get("in_total", 0.0)))],
            ["支出总额", self._fmt_amount(float(summary.get("out_total", 0.0)))],
            ["净额", self._fmt_amount(float(summary.get("net_amount", 0.0)))],
            ["总金额", self._fmt_amount(float(summary.get("total_amount", 0.0)))],
        ]
        th = float(result.extra.get("threshold", params.large_amount_threshold))
        if module_id in (AnalysisModuleId.LARGE_INOUT, AnalysisModuleId.LARGE_FLOW):
            summary_rows.append(["大额阈值", f"{th:.2f}"])
        if module_id == AnalysisModuleId.LARGE_INOUT:
            summary_rows.extend(
                [
                    ["大额收入笔数", str(int(result.extra.get("large_in_count", 0)))],
                    ["大额收入合计", self._fmt_amount(float(result.extra.get("large_in_total", 0.0)))],
                    ["大额支出笔数", str(int(result.extra.get("large_out_count", 0)))],
                    ["大额支出合计", self._fmt_amount(float(result.extra.get("large_out_total", 0.0)))],
                ]
            )
        if module_id in (AnalysisModuleId.SPECIAL_AMOUNT, AnalysisModuleId.SPECIAL_TIME):
            rc = result.extra.get("rule_hit_counts") or {}
            summary_rows.append(["规则命中统计", "; ".join(f"{k}({v}笔)" for k, v in list(rc.items())[:30])])

        stat_rows = [
            ["主体数量", str(summary.get("person_count", 0))],
            ["对手数量", str(summary.get("counterparty_count", 0))],
            ["分币种", self._query_service._format_currency_breakdown(summary.get("currency_breakdown", {}))],  # noqa: SLF001
            ["主要对手", self._query_service._format_top_counterparties(summary.get("top_counterparties", []))],  # noqa: SLF001
            ["时间分布", self._query_service._format_time_period_stats(summary.get("time_period_stats", {}))],  # noqa: SLF001
        ]

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            pd.DataFrame(detail_rows, columns=detail_headers).to_excel(writer, index=False, sheet_name="模块明细")
            pd.DataFrame(summary_rows, columns=["指标", "值"]).to_excel(writer, index=False, sheet_name="模块汇总")
            pd.DataFrame(stat_rows, columns=["统计项", "值"]).to_excel(writer, index=False, sheet_name="模块统计")

            if module_id == AnalysisModuleId.LARGE_FLOW:
                all_records = self._query_service.query_unified_records(import_batch_id, None)
                large = filter_large_transactions(all_records, th)
                flow_all = aggregate_large_flow(large, top_n=100_000)
                flow_headers = ["对手卡号", "对手名称", "流入", "流出", "净额", "笔数", "总发生额"]
                flow_data = [
                    [
                        str(r.get("counterparty_account") or ""),
                        str(r.get("counterparty_name") or ""),
                        self._fmt_amount(float(r.get("in_total", 0))),
                        self._fmt_amount(float(r.get("out_total", 0))),
                        self._fmt_amount(float(r.get("net", 0))),
                        str(int(r.get("count", 0))),
                        self._fmt_amount(float(r.get("volume", 0))),
                    ]
                    for r in flow_all
                ]
                pd.DataFrame(flow_data, columns=flow_headers).to_excel(
                    writer, index=False, sheet_name="按对手卡号汇总"
                )
                fws = writer.sheets.get("按对手卡号汇总")
                if fws is not None:
                    self._apply_sheet_style(fws)

            for name in ("模块明细", "模块汇总", "模块统计"):
                ws = writer.sheets.get(name)
                if ws is not None:
                    self._apply_sheet_style(ws)

        return str(out_file)

    def export_unified_record_table(
        self,
        output_path: str,
        records: list[dict[str, str]],
        *,
        sheet_name: str = "命中明细",
    ) -> str:
        """导出与整合明细字段一致的单表 xlsx（用于界面命中预览的佐证导出）。"""
        try:
            import pandas as pd
        except ImportError as err:
            raise RuntimeError("缺少 pandas，请先执行: pip install -r requirements.txt") from err

        detail_headers = [
            "数据来源",
            "银行类别",
            "姓名",
            "卡号",
            "时间戳",
            "收支标志",
            "币种",
            "金额",
            "余额",
            "对手名",
            "对手卡号",
            "交易描述",
            "备注",
        ]
        keys = (
            "data_source",
            "bank_type",
            "person_name",
            "acct_no",
            "txn_time",
            "txn_direction",
            "currency",
            "amount",
            "balance",
            "counterparty_name",
            "counterparty_account",
            "txn_desc",
            "remark",
        )
        detail_rows = [[row.get(k, "") for k in keys] for row in records]

        out_file = Path(output_path)
        if out_file.suffix.lower() != ".xlsx":
            out_file = out_file.with_suffix(".xlsx")
        out_file.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            pd.DataFrame(detail_rows, columns=detail_headers).to_excel(
                writer, index=False, sheet_name=sheet_name
            )
            ws = writer.sheets.get(sheet_name)
            if ws is not None:
                self._apply_sheet_style(ws)

        return str(out_file)

    def _build_name_match_diagnosis_sheet(self, import_batch_id: str) -> tuple[list[str], list[list[str]]]:
        """Diagnose why person_name cannot be filled for some txn rows."""
        headers = ["卡号", "卡号匹配Key(前6后4)", "是否开户表存在", "是否Key存在", "匹配姓名(若唯一)", "原因示例", "数据来源"]
        txn_rows = self._client.query_all(
            """
            SELECT acct_no, source_name
            FROM std_bank_txn
            WHERE import_batch_id=?
              AND (person_name IS NULL OR TRIM(person_name)='');
            """,
            (import_batch_id,),
        )
        if not txn_rows:
            return headers, []
        acc_rows = self._client.query_all(
            """
            SELECT acct_no, person_name
            FROM std_bank_account
            WHERE import_batch_id=?
              AND person_name IS NOT NULL AND TRIM(person_name) <> '';
            """,
            (import_batch_id,),
        )
        exact: dict[str, set[str]] = {}
        key_map: dict[str, set[str]] = {}

        def norm_acct(x: str) -> str:
            t = (x or "").strip().replace(" ", "")
            if t.endswith(".0") and t[:-2].isdigit():
                t = t[:-2]
            return t

        def acct_key(x: str) -> str:
            digits = "".join(ch for ch in norm_acct(x) if ch.isdigit())
            if len(digits) >= 10:
                return f"{digits[:6]}_{digits[-4:]}"
            return ""

        for acct_no, person_name in acc_rows:
            a = norm_acct("" if acct_no is None else str(acct_no))
            n = ("" if person_name is None else str(person_name)).strip()
            if not a or not n:
                continue
            exact.setdefault(a, set()).add(n)
            k = acct_key(a)
            if k:
                key_map.setdefault(k, set()).add(n)

        out: list[list[str]] = []
        seen = set()
        for acct_no, source_name in txn_rows:
            acct = norm_acct("" if acct_no is None else str(acct_no))
            if not acct:
                continue
            if acct in seen:
                continue
            seen.add(acct)
            k = acct_key(acct)
            exact_names = exact.get(acct, set())
            key_names = key_map.get(k, set()) if k else set()
            has_exact = "是" if acct in exact else "否"
            has_key = "是" if k and k in key_map else "否"
            match_name = ""
            reason = ""
            if exact_names:
                if len(exact_names) == 1:
                    match_name = next(iter(exact_names))
                    reason = "开户表精确匹配"
                else:
                    reason = "开户表精确匹配多姓名冲突"
            elif key_names:
                if len(key_names) == 1:
                    match_name = next(iter(key_names))
                    reason = "掩码Key匹配"
                else:
                    reason = "掩码Key匹配多姓名冲突"
            else:
                reason = "开户表未找到该卡号/Key"
            out.append([acct, k, has_exact, has_key, match_name, reason, "" if source_name is None else str(source_name)])
        return headers, out

    def export_filtered_summary(
        self,
        import_batch_id: str,
        output_path: str,
        filters: BankQueryFilters | None = None,
    ) -> str:
        """Export filtered details + generated description."""
        try:
            import pandas as pd
        except ImportError as err:
            raise RuntimeError("缺少 pandas，请先执行: pip install -r requirements.txt") from err
        filters = filters or BankQueryFilters()
        records = self._query_service.query_unified_records(import_batch_id, filters)
        params = ModuleParams()
        amount_counts = build_amount_rounded_counts(records)
        records = [
            {**row, "remark": build_custom_filter_analysis_remark(row, params, amount_counts)}
            for row in records
        ]
        summary = self._query_service.summarize(records)
        description = self._query_service.render_description(filters, summary)
        out_file = Path(output_path)
        if out_file.suffix.lower() != ".xlsx":
            out_file = out_file.with_suffix(".xlsx")
        out_file.parent.mkdir(parents=True, exist_ok=True)
        detail_headers = [
            "数据来源",
            "银行类别",
            "姓名",
            "卡号",
            "时间戳",
            "收支标志",
            "币种",
            "金额",
            "余额",
            "对手名",
            "对手卡号",
            "交易描述",
            "备注",
        ]
        detail_rows = [[row.get(k, "") for k in (
            "data_source",
            "bank_type",
            "person_name",
            "acct_no",
            "txn_time",
            "txn_direction",
            "currency",
            "amount",
            "balance",
            "counterparty_name",
            "counterparty_account",
            "txn_desc",
            "remark",
        )] for row in records]
        summary_rows = [
            ["交易笔数", str(summary.get("txn_count", 0))],
            ["收入总额", self._fmt_amount(float(summary.get("in_total", 0.0)))],
            ["支出总额", self._fmt_amount(float(summary.get("out_total", 0.0)))],
            ["净额", self._fmt_amount(float(summary.get("net_amount", 0.0)))],
            ["总金额", self._fmt_amount(float(summary.get("total_amount", 0.0)))],
            ["描述", description],
        ]
        stat_rows = [
            ["主体数量", str(summary.get("person_count", 0))],
            ["对手数量", str(summary.get("counterparty_count", 0))],
            ["分币种", self._query_service._format_currency_breakdown(summary.get("currency_breakdown", {}))],  # noqa: SLF001
            ["主要对手", self._query_service._format_top_counterparties(summary.get("top_counterparties", []))],  # noqa: SLF001
            [
                "对手集中度",
                (
                    f"Top1占比{float(summary.get('counterparty_concentration', {}).get('top1_ratio', 0.0))*100:.2f}%"
                    f"，Top3占比{float(summary.get('counterparty_concentration', {}).get('top3_ratio', 0.0))*100:.2f}%"
                ),
            ],
            ["时间分布", self._query_service._format_time_period_stats(summary.get("time_period_stats", {}))],  # noqa: SLF001
            ["异常标签", self._query_service._format_remark_tag_stats(summary.get("remark_tag_stats", []))],  # noqa: SLF001
        ]
        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            pd.DataFrame(detail_rows, columns=detail_headers).to_excel(writer, index=False, sheet_name="筛选结果")
            pd.DataFrame(summary_rows, columns=["指标", "值"]).to_excel(writer, index=False, sheet_name="筛选描述")
            pd.DataFrame(stat_rows, columns=["统计项", "值"]).to_excel(writer, index=False, sheet_name="筛选统计")
            dws = writer.sheets.get("筛选结果")
            sws = writer.sheets.get("筛选描述")
            tws = writer.sheets.get("筛选统计")
            if dws is not None:
                self._apply_sheet_style(dws)
            if sws is not None:
                self._apply_sheet_style(sws)
            if tws is not None:
                self._apply_sheet_style(tws)
        return str(out_file)

    def _load_batch_all_raw_fields(self, import_batch_id: str) -> tuple[list[str], list[list[str]]]:
        """优先导出标准层总体表；若无标准层数据则回退 raw 并集导出。"""
        std_headers, std_rows = self._load_std_overview_rows(import_batch_id)
        if std_rows:
            return std_headers, std_rows
        return self._load_raw_union_rows(import_batch_id)

    def _load_std_overview_rows(self, import_batch_id: str) -> tuple[list[str], list[list[str]]]:
        """Load standardized overview rows for bank flow."""
        headers = [
            "数据来源",
            "银行类别",
            "姓名",
            "卡号",
            "时间戳",
            "收支标志",
            "币种",
            "金额",
            "余额",
            "对手名",
            "对手卡号",
            "交易描述",
            "备注",
        ]
        records = self._query_service.query_unified_records(import_batch_id, BankQueryFilters())
        if not records:
            return headers, []
        out_rows: list[list[str]] = []
        for rec in records:
            source_name = self._sanitize_export_cell(rec["data_source"])
            bank_name = self._sanitize_export_cell(rec["bank_type"])
            person_name = self._sanitize_export_cell(rec["person_name"])
            acct_no = self._sanitize_export_cell(rec["acct_no"])
            when = self._sanitize_export_cell(rec["txn_time"])
            direction = self._sanitize_export_cell(rec["txn_direction"])
            currency = self._sanitize_export_cell(rec["currency"])
            amount = rec["amount"]
            balance = rec["balance"]
            cp_name = self._sanitize_export_cell(rec["counterparty_name"])
            cp_acct = self._sanitize_export_cell(rec["counterparty_account"])
            summary = self._sanitize_export_cell(rec["txn_desc"])
            remark = self._sanitize_export_cell(rec["remark"])
            out_rows.append(
                [source_name, bank_name, person_name, acct_no, when, direction, currency, amount, balance, cp_name, cp_acct, summary, remark]
            )
        return headers, out_rows

    def _load_std_records(self, import_batch_id: str) -> list[dict[str, str]]:
        """Load standardized rows for downstream detail/stat sheets."""
        rows = self._client.query_all(
            """
            SELECT bank_name, source_name, person_name, acct_no, counterparty_name, counterparty_account,
                   txn_amount, txn_time, txn_direction, summary, remark, txn_org_no, txn_org_name
            FROM std_bank_txn
            WHERE import_batch_id=?
            ORDER BY source_name, std_id;
            """,
            (import_batch_id,),
        )
        output: list[dict[str, str]] = []
        for row in rows:
            output.append(
                {
                    "bank_name": "" if row[0] is None else str(row[0]),
                    "source_name": "" if row[1] is None else str(row[1]),
                    "person_name": "" if row[2] is None else str(row[2]),
                    "acct_no": "" if row[3] is None else str(row[3]),
                    "counterparty_name": "" if row[4] is None else str(row[4]),
                    "counterparty_account": "" if row[5] is None else str(row[5]),
                    "txn_amount": "" if row[6] is None else str(row[6]),
                    "txn_time": "" if row[7] is None else str(row[7]),
                    "txn_direction": "" if row[8] is None else str(row[8]),
                    "summary": "" if row[9] is None else str(row[9]),
                    "remark": "" if row[10] is None else str(row[10]),
                    "txn_org_no": "" if row[11] is None else str(row[11]),
                    "txn_org_name": "" if row[12] is None else str(row[12]),
                }
            )
        return output

    def _build_counterparty_detail_sheet(self, records: list[dict[str, str]]) -> tuple[list[str], list[list[str]]]:
        """Build A->B / B->A / gap detail rows."""
        headers = [
            "银行",
            "本方姓名",
            "本方账号",
            "对手名",
            "对手账号",
            "A->B总额",
            "B->A总额",
            "双方往来差额",
            "最近时间点",
            "摘要示例",
            "备注",
        ]
        groups: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
        for rec in records:
            key = (
                rec["bank_name"],
                rec["person_name"],
                rec["acct_no"],
                rec["counterparty_name"],
                rec["counterparty_account"],
            )
            g = groups.setdefault(
                key, {"a2b": 0.0, "latest_time": "", "summary": "", "remark": "", "counter_key": None}
            )
            amount = self._safe_float(rec["txn_amount"])
            direction = rec["txn_direction"]
            if direction == "支出":
                g["a2b"] += abs(amount)
            elif direction == "收入":
                g["a2b"] -= abs(amount)
            if rec["txn_time"] and rec["txn_time"] > g["latest_time"]:
                g["latest_time"] = rec["txn_time"]
            if not g["summary"] and rec["summary"]:
                g["summary"] = rec["summary"]
            g["remark"] = self._merge_tag_text(g["remark"], rec["remark"])
            g["counter_key"] = (
                rec["bank_name"],
                rec["counterparty_name"],
                rec["counterparty_account"],
                rec["person_name"],
                rec["acct_no"],
            )

        out_rows: list[list[str]] = []
        for key, g in groups.items():
            reverse = groups.get(g["counter_key"], {})
            b2a = float(reverse.get("a2b", 0.0))
            a2b = float(g["a2b"])
            gap = a2b - b2a
            out_rows.append(
                [
                    key[0],
                    key[1],
                    key[2],
                    key[3],
                    key[4],
                    self._fmt_amount(a2b),
                    self._fmt_amount(b2a),
                    self._fmt_amount(gap),
                    g["latest_time"],
                    g["summary"],
                    g["remark"],
                ]
            )
        out_rows.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4]))
        return headers, out_rows

    def _build_person_bank_detail_sheet(self, records: list[dict[str, str]]) -> tuple[list[str], list[list[str]]]:
        """Build person-bank transaction detail rows."""
        headers = [
            "银行",
            "姓名",
            "账号",
            "对手名",
            "对手账号",
            "发生金额",
            "发生时间",
            "借贷方向",
            "交易行号",
            "交易行名",
            "摘要",
            "备注",
            "数据来源",
        ]
        rows = [
            [
                rec["bank_name"],
                rec["person_name"],
                rec["acct_no"],
                rec["counterparty_name"],
                rec["counterparty_account"],
                self._fmt_amount(self._safe_float(rec["txn_amount"])),
                rec["txn_time"],
                rec["txn_direction"],
                rec["txn_org_no"],
                rec["txn_org_name"],
                rec["summary"],
                rec["remark"],
                rec["source_name"],
            ]
            for rec in records
        ]
        rows.sort(key=lambda x: (x[0], x[1], x[2], x[6], x[3], x[4]))
        return headers, rows

    def _build_person_bank_stat_sheet(self, records: list[dict[str, str]]) -> tuple[list[str], list[list[str]]]:
        """Build person-bank aggregated stats."""
        headers = ["银行", "姓名", "账号", "收入总额", "支出总额", "净额", "交易笔数", "备注"]
        groups: dict[tuple[str, str, str], dict[str, Any]] = {}
        for rec in records:
            key = (rec["bank_name"], rec["person_name"], rec["acct_no"])
            g = groups.setdefault(key, {"in_total": 0.0, "out_total": 0.0, "count": 0, "remark": ""})
            amount = abs(self._safe_float(rec["txn_amount"]))
            if rec["txn_direction"] == "收入":
                g["in_total"] += amount
            elif rec["txn_direction"] == "支出":
                g["out_total"] += amount
            g["count"] += 1
            g["remark"] = self._merge_tag_text(g["remark"], rec["remark"])
        rows: list[list[str]] = []
        for key, g in groups.items():
            net = g["in_total"] - g["out_total"]
            rows.append(
                [
                    key[0],
                    key[1],
                    key[2],
                    self._fmt_amount(g["in_total"]),
                    self._fmt_amount(g["out_total"]),
                    self._fmt_amount(net),
                    str(g["count"]),
                    g["remark"],
                ]
            )
        rows.sort(key=lambda x: (x[0], x[1], x[2]))
        return headers, rows

    def _load_raw_union_rows(self, import_batch_id: str) -> tuple[list[str], list[list[str]]]:
        """按批次汇总所有 raw 表的 src_* 字段，输出统一宽表。"""
        file_rows = self._client.query_all(
            """
            SELECT file_id, file_name
            FROM meta_bank_files
            WHERE import_batch_id=?;
            """,
            (import_batch_id,),
        )
        file_name_map = {int(row[0]): str(row[1]) for row in file_rows if row and row[0] is not None}

        sheets = self._client.query_all(
            """
            SELECT DISTINCT raw_table_name
            FROM meta_bank_sheets s
            JOIN meta_bank_files f ON f.file_id=s.file_id
            WHERE f.import_batch_id=?
            ORDER BY raw_table_name;
            """,
            (import_batch_id,),
        )
        table_names = [str(row[0]) for row in sheets]
        if not table_names:
            return ["提示"], [["当前批次无可导出的原始字段数据"]]

        all_src_cols: list[str] = []
        for table in table_names:
            info = self._client.query_all(f"PRAGMA table_info({self._client.quote_ident(table)});")
            cols = [str(x[1]) for x in info if str(x[1]).startswith("src_")]
            for col in cols:
                if col not in all_src_cols:
                    all_src_cols.append(col)

        if not all_src_cols:
            return ["提示"], [["当前批次原始表未识别到 src_ 字段"]]

        headers = ["数据来源"] + [self._display_src_name(c) for c in all_src_cols]
        output_rows: list[list[str]] = []
        for table in table_names:
            info = self._client.query_all(f"PRAGMA table_info({self._client.quote_ident(table)});")
            cols = [str(x[1]) for x in info if str(x[1]).startswith("src_")]
            if not cols:
                continue
            sql_cols = ", ".join([self._client.quote_ident(c) for c in cols])
            rows = self._client.query_all(
                f"""
                SELECT source_file_id, source_sheet, {sql_cols}
                FROM {self._client.quote_ident(table)}
                WHERE import_batch_id=?
                ORDER BY raw_id;
                """,
                (import_batch_id,),
            )
            col_idx = {name: idx for idx, name in enumerate(cols, start=2)}
            for row in rows:
                file_id = int(row[0]) if row[0] is not None else 0
                file_name = file_name_map.get(file_id, "")
                source_sheet = "" if row[1] is None else str(row[1])
                source_name = self._build_row_source_name(file_name, source_sheet, fallback_table=table)
                line: list[str] = [self._sanitize_export_cell(source_name)]
                for c in all_src_cols:
                    idx = col_idx.get(c)
                    if idx is None:
                        line.append("")
                    else:
                        val: Any = row[idx]
                        text = "" if val is None else str(val)
                        line.append(self._sanitize_export_cell(text))
                output_rows.append(line)
        return headers, output_rows

    @staticmethod
    def _display_src_name(sql_name: str) -> str:
        return sql_name[4:] if sql_name.startswith("src_") else sql_name

    @staticmethod
    def _build_row_source_name(file_name: str, sheet_name: str, fallback_table: str) -> str:
        """逐行来源名：原文件名(去扩展) + 工作表名；缺失时回退到表名。"""
        stem = Path(file_name).stem.strip() if file_name else ""
        sheet = str(sheet_name).strip() if sheet_name else ""
        if stem and sheet:
            return f"{stem}_{sheet}"
        if stem:
            return stem
        if sheet:
            return sheet
        text = fallback_table
        if text.startswith("raw_"):
            text = text[4:]
        return text.replace("_", " ")

    def export_batch_to_csv(self, import_batch_id: str, output_path: str) -> str:
        """兼容旧调用：输出同口径“全字段合并”CSV。"""
        headers, rows = self._load_batch_all_raw_fields(import_batch_id)
        out_file = Path(output_path)
        out_file.parent.mkdir(parents=True, exist_ok=True)

        lines = [",".join(headers)]
        for row in rows:
            values = [self._csv_escape("" if value is None else str(value)) for value in row]
            lines.append(",".join(values))
        out_file.write_text("\n".join(lines), encoding="utf-8-sig")
        return str(out_file)

    def _csv_escape(self, value: str) -> str:
        """Escape one csv value."""
        escaped = value.replace('"', '""')
        return f"\"{escaped}\""

    def _sanitize_export_cell(self, value: str) -> str:
        """Remove noisy bank disclaimers from exported cells."""
        text = (value or "").strip()
        if not text:
            return ""
        if text.startswith("数据截至") and "生产系统为准" in text:
            return ""
        if all(keyword in text for keyword in self.DISCLAIMER_KEYWORDS):
            return ""
        return text

    def _apply_sheet_style(self, ws) -> None:
        """Apply borders, alignment, freeze header row, and autosize columns."""
        from openpyxl.styles import Alignment, Border, Side

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(vertical="center", wrap_text=True)

        # Freeze first row.
        ws.freeze_panes = "A2"

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = align
                cell.border = border

        # Auto-fit each column width based on displayed content length.
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                text = "" if cell.value is None else str(cell.value)
                if len(text) > max_len:
                    max_len = len(text)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 80)

    def _merge_same_source_cells(self, ws) -> None:
        """Merge contiguous rows that share the same source value in column A."""
        if ws.max_row <= 2:
            return

        start = 2
        current = ws.cell(row=2, column=1).value
        for row_idx in range(3, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=1).value
            if value != current:
                if current not in (None, "") and row_idx - 1 > start:
                    ws.merge_cells(start_row=start, start_column=1, end_row=row_idx - 1, end_column=1)
                start = row_idx
                current = value

        if current not in (None, "") and ws.max_row > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=ws.max_row, end_column=1)

    def _safe_float(self, value: str) -> float:
        text = (value or "").strip().replace(",", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0

    def _fmt_amount(self, value: float) -> str:
        return f"{value:.2f}"

    def _merge_tag_text(self, left: str, right: str) -> str:
        items: list[str] = []
        for src in (left, right):
            if not src:
                continue
            for token in [x.strip() for x in src.split(";") if x.strip()]:
                if token not in items:
                    items.append(token)
        return "; ".join(items)

